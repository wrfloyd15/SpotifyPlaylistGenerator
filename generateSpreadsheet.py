import spotipy
import spotipy.util as util
from client_info import my_username, my_client_id, my_client_secret,scope
from openpyxl import Workbook
from openpyxl.styles import Font

token = util.prompt_for_user_token(my_username, scope, client_id=my_client_id,client_secret=my_client_secret,redirect_uri='http://localhost')

def add_to_spreadsheet(item,sheet,currentRow):

    # get info to add to spreadsheet
    date_added = item['added_at']
    date_added = date_added[0:10]
    item = item['track']
    title = item['name']
    artist = item['album']['artists'][0]['name']
    release_date = item['album']['release_date']
    url = item['href']
    popularity = item['popularity']

    # add info to spreadsheet
    sheet.cell(row=currentRow, column=1, value=title)
    sheet.cell(row=currentRow, column=2, value=artist)
    sheet.cell(row=currentRow, column=3, value=release_date)
    sheet.cell(row=currentRow, column=4, value=date_added)
    sheet.cell(row=currentRow, column=5, value=url)
    sheet.cell(row=currentRow, column=6, value=popularity)

def initialize_spreadsheet(sheet):
    # initialize spreadsheet
    sheet.cell(row=1, column=1, value='Title')
    sheet['A1'].font = Font(bold = True)
    sheet.column_dimensions['A'].width = 50
    sheet.cell(row=1, column=2, value='Artist')
    sheet['B1'].font = Font(bold = True)
    sheet.column_dimensions['B'].width = 20
    sheet.cell(row=1, column=3, value='Release Date')
    sheet['C1'].font = Font(bold = True)
    sheet.column_dimensions['C'].width = 12
    sheet.cell(row=1, column=4, value='Date Added')
    sheet['D1'].font = Font(bold = True)
    sheet.column_dimensions['D'].width = 12
    sheet.cell(row=1, column=5, value='URL')
    sheet['E1'].font = Font(bold = True)
    sheet.column_dimensions['E'].width = 53
    sheet.cell(row=1, column=6, value='Popularity')
    sheet['F1'].font = Font(bold = True)
    sheet.column_dimensions['F'].width = 12


# make sure the file ends with .xlsx
fileName = input("Enter a name for the spreadsheet: ")

if fileName.endswith(".xlsx") == False:
    fileName = fileName + ".xlsx"

# Create the spreadsheet if the token works
if token:
    sp = spotipy.Spotify(auth=token)

    # initialize spreadsheet
    print("Creating spreadsheet...")
    spreadsheet = Workbook()
    sheet = spreadsheet.active
    initialize_spreadsheet(sheet)

    # loop through saved songs and add them to spreadsheet
    print("Adding songs to spreadsheet...")
    row = 2
    trackCount=50
    i=0
    while (trackCount == 50):
        trackCount = 0
        results = sp.current_user_saved_tracks(limit = 50, offset = 50*i)
        for item in results['items']:
            add_to_spreadsheet(item,sheet,row)
            row+=1
            trackCount+=1
        i+=1

    # save the spreadsheet
    spreadsheet.save(fileName)
    songCount = 50*(i-1)+trackCount
    print (f"Added {songCount} songs to spreadsheet {fileName}")

else:
    print("Can't get token for", my_username)
